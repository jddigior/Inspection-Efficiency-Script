"""
S&C Electric Canada
Developed by Jonathan DiGiorgio
July 2023 - August 2023
Python 3.11

Dependancies: 
- Openpyxl (version 3.1.2)
- Excel (Developed on Microsoft® Excel® for Microsoft 365 MSO Version 2305 Build 16.0.16501.20074 64-bit )

IT IS UNKNOWN IF THIS PROGRAM WILL FUNCTION ON UPDATED VERSIONS OF DEPENDANCIES
"""

# Modules and Libraries
import webbrowser
import time
from openpyxl import load_workbook
from datetime import date
import cv2
import os

# Variables
FILE_NAME = date.today().strftime('Small Packaging Inspection %Y.xlsx') # ADD WAY TO CHANGE THIS BASED ON YEAR
TEMP_FILE_NAME = 'Program-Files/temp.xlsx'
LINK_TEMPLATE = 'https://sourceone.sandc.ws/apps/drawingsearch?query='
PIC_LINK = r'file:///\\TOR-FS01.sandc.ws\QltyAssr\Inspections\Small Packaging Inspection' + '\\'

# Tries to load spreadsheet with exception handling for when another team member is using the file
def canLoadXl():
    try:
        wb = load_workbook(FILE_NAME, data_only = True)
        wb.save(FILE_NAME)
    except PermissionError:
        try:
            wb = load_workbook(TEMP_FILE_NAME, data_only = True)
            print('Someone has the file open, you\'re data will be saved and uploaded later.')
        except PermissionError:
            print('Someone already has the document open. Try again later.')
            return False, 'Error'
        else:
            return False, wb
    else:
        return True, wb
    
# Calls above function to attempt file load until sucessful or exited
def loadXl():
    isLoaded, wb = canLoadXl()
    while (wb == 'Error'):
        input('Press Enter to try again.')
        isLoaded, wb = canLoadXl()

    return wb, isLoaded

#Finds the next empty row in the catalog number column
def nextFreeRow(sheet):
    count = 0
    rowFound = False
    while not rowFound:
        count += 1
        if sheet.cell(row = count, column = 3).value == None: # Catalog num are column 3
            rowFound = True
    return count

#checks if a date was already listed in the spreadsheet
def dateIsListed(count):

    while (count > 0):
        #Checks the two most common date formats from the sheet, (00:00:00 is added by excel but not shown in the cell)
        if str(ws.cell(row = count, column = 1).value) in {date.today().strftime("%d/%m/%Y"), date.today().strftime("%Y-%m-%d 00:00:00")}:
            return True
        elif ws.cell(row = count, column = 1).value is not None:
            return False
        count -= 1
    else:
        return False

# Uses sample size chart standard for the following numbers
def numToInspect(qty):
    qty = int(qty)

    if qty <= 2:
        return qty
    elif qty <= 25:
        return 2
    elif qty <= 50:
        return 3
    elif qty <= 90:
        return 5
    elif qty <= 150:
        return 8
    elif qty <= 280:
        return 13
    elif qty <= 500:
        return 20
    elif qty <= 1200:
        return 32
    elif qty <= 3200:
        return 50
    elif qty <= 10000:
        return 80
    elif qty <= 35000:
        return 125  #Standard ends here
    else:
        #return int((125/35000)*qty) # Linearization from previous 
        return int(0.9*(qty**0.474)) # Power trend line from Excel, R^2 = 0.987

# Opens drawing in browser
def openDrawing(catalogNum):
    link = LINK_TEMPLATE + catalogNum
    webbrowser.open(link)

# Initializes sheet based on month
def initSheet(wb):
    sheets = wb.sheetnames
    month = date.today().month # Saves month as a number 1-12
    ws = wb[sheets[month - 1]]
    return ws

# Attempts to move data from the temp file to the main
def uploadTemp(writeRow):
    wbTemp = load_workbook(TEMP_FILE_NAME, data_only = True)
    wsTemp = wbTemp.active #temp only has one sheet

    row = nextFreeRow(wsTemp)

    for i in range(1, row):
        if not dateIsListed(writeRow - 1): 
            ws['A' + str(writeRow)] = date.today().strftime("%d/%m/%Y")

        ws['B' + str(writeRow)] = wsTemp['B' + str(i)].value
        ws['C' + str(writeRow)] = wsTemp['C' + str(i)].value
        ws['D' + str(writeRow)] = wsTemp['D' + str(i)].value
        ws['E' + str(writeRow)] = wsTemp['E' + str(i)].value
        ws['F' + str(writeRow)] = wsTemp['F' + str(i)].value
        ws['G' + str(writeRow)] = wsTemp['G' + str(i)].value

        if wsTemp['H' + str(i)].value not in {None, ' '}:
            ws['H' + str(writeRow)] = wsTemp['H' + str(i)].value

        if wsTemp['I' + str(i)].value not in {None, ' '}:
            ws['I' + str(writeRow)].value = wsTemp['I' + str(i)].value
            ws['I' + str(writeRow)].hyperlink = wsTemp['I' + str(i)].hyperlink

        writeRow += 1

    wsTemp.delete_rows(1,row)
    wbTemp.save(TEMP_FILE_NAME)
    return writeRow

# Reads announcment file and prints it
def displayAnnouncments():
    file = open('Program-Files/QA-reminders.txt','r')

    reminders = file.read()
    print('---------------------------------------------------------------------------')
    print('REMINDERS: \n')
    print(reminders)
    print('---------------------------------------------------------------------------')

    file.close()

# Has user take a picture of the lot
def getLotPic(SO, catNum):
    cam = cv2.VideoCapture(0)

    while True:
        ret, frame = cam.read()
        if not ret:
            print("failed to grab frame")
            ret = 'Empty'
            break

        cv2.imshow("Click space to take a picture, escape to exit", frame)

        k = cv2.waitKey(1)
        if k%256 == 27:
            # ESC pressed
            print("Escape hit, closing...")
            ret = 'Empty'
            break
        elif k%256 == 32:
            # SPACE pressed
            path = 'Program-Files/images/' + str(SO)
            isExist = os.path.exists(path)
            if not isExist:
                os.makedirs(path)
            
            count = 0
            path = 'Program-Files/images/' + str(SO) + "/" + str(catNum) + ".png"
            while os.path.exists(path):
                count += 1
                path = 'Program-Files/images/' + str(SO) + "/" + str(catNum) + '-(' + str(count) + ')' + ".png"

            cv2.imwrite(path, frame)
            print("{}.png written!".format(str(catNum)))

            ret = path
            break

    cam.release()
    cv2.destroyAllWindows()
    return ret

## ------------------ MAIN ------------------ ## 

displayAnnouncments()

# Check to see if xlsx is open already and upload temp contents if it isn't
wb, isLoaded = loadXl()
if(isLoaded == True):
    ws = initSheet(wb)
    writeRow = nextFreeRow(ws)
    writeRow = uploadTemp(writeRow) #upload data from temp to the main file
    wb.save(FILE_NAME)
else:
    wb.save(TEMP_FILE_NAME)

user = input('Please enter your name (LASTNAME, FIRSTNAME): ')

while True: #Continues until user is done inspecting

    # SALES ORDER INPUT
    SO = input('Enter Sales Order (Check Oracle): ')
    while (len(SO) != 6) or (not SO.isdigit()): #Sales orders must be 6 digit numbers
        SO = input('Error. Please enter a 6 digit Sales Order: ')

    # CATALOG NUMBER INPUT
    catNum = input('Enter Catalog Number: ')

    # LOT QUANTITY INPUT
    lotQty = input('Enter lot quantity: ')
    while not lotQty.isdigit():
        lotQty = input('Error! Please enter a number quantity : ')

    # SAMPLE SIZE INSPECTION STANDARD
    inspQty = numToInspect(lotQty)
    print("Please inspect {} part(s)".format(inspQty))
    time.sleep(1.5)

     # OPEN RESPECTIVE DRAWING
    openDrawing(catNum)

    # INSPECTION RESULTS INPUT
    result = input('Pass or Fail? : ')
    while result not in {'Pass','pass','Fail','fail','p','P','f','F'}:
        result = input('Invalid input, try again: ')

    if result in {'p','P','pass'}:
        result = 'Pass'
    elif result in {'f','F','fail'}:
        result = 'Fail'

    # NOTE INPUT
    note = input('Enter a note (Press enter to skip): ')

    #Prompt to take a picture of lot
    print('Press space to take a picture of the lot (esc to skip)')
    path = getLotPic(SO, catNum)

    # Open sheet again to quickly add data
    wb, isLoaded = loadXl()
    if(isLoaded == True):
        ws = initSheet(wb)
    else:
        ws = wb.active
    
    writeRow = nextFreeRow(ws)

    # Add date if needed
    if not dateIsListed(writeRow - 1):
        ws['A' + str(writeRow)] = date.today().strftime("%d/%m/%Y")

    # Add everything else
    ws['B' + str(writeRow)] = int(SO)
    ws['C' + str(writeRow)] = catNum
    ws['D' + str(writeRow)] = int(inspQty)
    ws['E' + str(writeRow)] = int(lotQty)
    ws['F' + str(writeRow)] = result
    ws['G' + str(writeRow)] = user
    if note not in {None, ' '}:
        ws['H' + str(writeRow)] = note
    if path != 'Empty':
        ws['I' + str(writeRow)].value = "picture"
        ws['I' + str(writeRow)].hyperlink = PIC_LINK + path

    # Save data
    if(isLoaded == True):
        wb.save(FILE_NAME)
    else:
        wb.save(TEMP_FILE_NAME)
    print('-----------------------------------Data saved!----------------------------------------')

    # Break the loop if user does not want to enter another inspection
    anotherInsp = input('Click enter to do another inspection. Enter \'q\' to quit. ')
    if anotherInsp in {'q','Q'}:
        break

input('---------------------------------------------------------------------------\nPRESS ENTER TO CLOSE...')
